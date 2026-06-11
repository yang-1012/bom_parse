"""Excel 导出服务"""

import logging
import os
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.models.device import Device

logger = logging.getLogger("ParseApp")

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
CELL_FONT = Font(name="微软雅黑", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

DETAIL_HEADERS = [
    "器件编码", "器件描述", "位号", "数量",
    "封装", "管脚数", "分类", "T面数量", "B面数量",
    "总焊点数", "折算后件数",
]


def export_to_excel(
    devices: list[Device],
    source_path: str,
    output_dir: Optional[str] = None,
) -> str:
    """导出 Excel 报表

    文件命名: {源文件名}_解析.xlsx
    包含两个 Sheet: 分类明细 + 统计汇总
    """
    base_name = os.path.splitext(os.path.basename(source_path))[0]
    output_dir = output_dir or os.path.dirname(source_path)
    output_path = os.path.join(output_dir, f"{base_name}_解析.xlsx")

    wb = openpyxl.Workbook()
    # 移除默认空白 sheet
    wb.remove(wb.active)

    _write_detail_sheet(wb, devices, "全部分类")
    _write_summary_sheet(wb, devices)

    # 按分类分组，每个分类一个独立 sheet
    from collections import defaultdict
    cats = defaultdict(list)
    for dev in devices:
        cats[dev.classification].append(dev)
    for cat_name in sorted(cats):
        sheet_name = cat_name[:31]
        _write_detail_sheet(wb, cats[cat_name], sheet_name)

    wb.save(output_path)
    logger.info(f"Excel 导出完成: {output_path}")
    return output_path


def _write_detail_sheet(wb, devices: list[Device], title: str = "分类明细") -> None:
    ws = wb.create_sheet(title)

    # 表头
    for col_idx, header in enumerate(DETAIL_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER

    # 数据
    for row_idx, dev in enumerate(devices, 2):
        values = [
            dev.code, dev.description, dev.refdes, dev.quantity,
            dev.package, dev.pin_count, dev.classification,
            dev.t_side_count, dev.b_side_count,
            dev.total_pads, round(dev.converted_qty, 2),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER

    ws.auto_filter.ref = ws.dimensions


def _write_summary_sheet(wb, devices: list[Device]) -> None:
    ws = wb.create_sheet("统计汇总")

    # 按分类统计
    from collections import defaultdict
    cat_types = defaultdict(int)
    cat_qty = defaultdict(int)
    cat_t = defaultdict(int)
    cat_b = defaultdict(int)
    cat_qtys = defaultdict(float)
    for dev in devices:
        cat_types[dev.classification] += 1
        cat_qty[dev.classification] += dev.quantity
        cat_t[dev.classification] += dev.t_side_count
        cat_b[dev.classification] += dev.b_side_count
        cat_qtys[dev.classification] += dev.converted_qty

    headers = ["分类", "器件行数", "总数量", "T面件数", "B面件数", "折算后件数"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = CENTER

    row = 2
    for cat in sorted(cat_types.keys()):
        ws.cell(row=row, column=1, value=cat).font = CELL_FONT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=2, value=cat_types[cat]).font = CELL_FONT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = CENTER
        ws.cell(row=row, column=3, value=cat_qty[cat]).font = CELL_FONT
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = CENTER
        ws.cell(row=row, column=4, value=cat_t[cat]).font = CELL_FONT
        ws.cell(row=row, column=4).border = THIN_BORDER
        ws.cell(row=row, column=4).alignment = CENTER
        ws.cell(row=row, column=5, value=cat_b[cat]).font = CELL_FONT
        ws.cell(row=row, column=5).border = THIN_BORDER
        ws.cell(row=row, column=5).alignment = CENTER
        ws.cell(row=row, column=6, value=round(cat_qtys[cat], 2)).font = CELL_FONT
        ws.cell(row=row, column=6).border = THIN_BORDER
        ws.cell(row=row, column=6).alignment = CENTER
        row += 1

    # 合计行
    total_font = Font(name="微软雅黑", size=10, bold=True)
    ws.cell(row=row, column=1, value="合计").font = total_font
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2, value=sum(cat_types.values())).font = total_font
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.cell(row=row, column=2).alignment = CENTER
    ws.cell(row=row, column=3, value=sum(cat_qty.values())).font = total_font
    ws.cell(row=row, column=3).border = THIN_BORDER
    ws.cell(row=row, column=3).alignment = CENTER
    ws.cell(row=row, column=4, value=sum(cat_t.values())).font = total_font
    ws.cell(row=row, column=4).border = THIN_BORDER
    ws.cell(row=row, column=4).alignment = CENTER
    ws.cell(row=row, column=5, value=sum(cat_b.values())).font = total_font
    ws.cell(row=row, column=5).border = THIN_BORDER
    ws.cell(row=row, column=5).alignment = CENTER
    ws.cell(row=row, column=6, value=round(sum(cat_qtys.values()), 2)).font = total_font
    ws.cell(row=row, column=6).border = THIN_BORDER
    ws.cell(row=row, column=6).alignment = CENTER

    # 使用默认列宽
